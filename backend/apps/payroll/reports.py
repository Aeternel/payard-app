# ruff: noqa: E501

import io
from dataclasses import dataclass
from decimal import Decimal

from django.utils.html import escape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import PayrollCycle
from .services import money

GREEN = "165C43"
DARK_GREEN = "0D3D2C"
LIME = "CCF06A"
PAPER = "F4F0E7"
LIGHT_GREEN = "D9EFDF"
LIGHT_AMBER = "FFF0C8"
LINE = "D9DDD5"


@dataclass(frozen=True)
class PayrollReport:
    cycle: PayrollCycle
    lines: list
    totals: dict


def report_data(cycle):
    lines = list(
        cycle.lines.select_related("worker", "manual_override_by").order_by(
            "worker__worker_code"
        )
    )
    totals = {
        "worker_count": len(lines),
        "contract_basic": sum((line.contract_basic for line in lines), Decimal("0")),
        "regular_pay": sum((line.regular_pay for line in lines), Decimal("0")),
        "overtime_pay": sum((line.overtime_pay for line in lines), Decimal("0")),
        "allowances": sum((line.allowances for line in lines), Decimal("0")),
        "other_earnings": sum((line.other_earnings for line in lines), Decimal("0")),
        "gross_pay": sum((line.gross_pay for line in lines), Decimal("0")),
        "absence_deductions": sum(
            (line.absence_deductions for line in lines), Decimal("0")
        ),
        "advance_deductions": sum(
            (line.advance_deductions for line in lines), Decimal("0")
        ),
        "other_deductions": sum(
            (line.other_deductions for line in lines), Decimal("0")
        ),
        "calculated_net_pay": sum(
            (line.calculated_net_pay for line in lines), Decimal("0")
        ),
        "final_net_pay": sum((line.net_pay for line in lines), Decimal("0")),
        "manual_override_count": sum(line.manual_net_pay is not None for line in lines),
    }
    return PayrollReport(
        cycle=cycle,
        lines=lines,
        totals={key: money(value) if isinstance(value, Decimal) else value for key, value in totals.items()},
    )


def _aed(value):
    return f"AED {money(value):,.2f}"


def _filename(report, extension):
    safe_name = "".join(
        character.lower() if character.isalnum() else "-"
        for character in report.cycle.name
    ).strip("-")
    return f"payyard-payroll-{safe_name}-v{report.cycle.version}.{extension}"


def build_html_report(report):
    status_note = (
        "FINAL PAYROLL REPORT"
        if report.cycle.status
        in {
            PayrollCycle.Status.APPROVED,
            PayrollCycle.Status.LOCKED,
            PayrollCycle.Status.EXPORTED,
            PayrollCycle.Status.PAID,
        }
        else "REVIEW REPORT - NOT YET FINANCE APPROVED"
    )
    rows = []
    for index, line in enumerate(report.lines, start=1):
        override = (
            f"<span class='override'>Override</span><br>{escape(line.manual_override_reason)}"
            if line.manual_net_pay is not None
            else ""
        )
        deductions = line.absence_deductions + line.advance_deductions + line.other_deductions
        rows.append(
            "<tr>"
            f"<td>{index}</td>"
            f"<td><strong>{escape(line.worker.worker_code)}</strong></td>"
            f"<td><strong>{escape(line.worker.full_name)}</strong>"
            f"<br><small>{escape(line.worker.job_title or 'Role not set')}</small></td>"
            f"<td>{_aed(line.contract_basic)}</td>"
            f"<td>{_aed(line.regular_pay)}</td>"
            f"<td>{_aed(line.overtime_pay)}</td>"
            f"<td>{_aed(line.allowances + line.other_earnings)}</td>"
            f"<td>{_aed(deductions)}</td>"
            f"<td>{_aed(line.calculated_net_pay)}</td>"
            f"<td><strong>{_aed(line.net_pay)}</strong><br><small>{override}</small></td>"
            "</tr>"
        )
    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{escape(report.cycle.name)} Payroll Report</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; padding: 32px; color: #12251d; background: #f4f0e7;
    font-family: Inter, Arial, sans-serif; }}
  .report {{ max-width: 1500px; margin: auto; background: #fffdf8; border: 1px solid #d9ddd5;
    border-radius: 18px; padding: 30px; }}
  header {{ display: flex; justify-content: space-between; gap: 24px; align-items: flex-start;
    border-bottom: 3px solid #165c43; padding-bottom: 20px; }}
  .brand {{ font-size: 24px; font-weight: 700; color: #0d3d2c; }}
  h1 {{ margin: 14px 0 6px; font-size: 36px; letter-spacing: -1px; }}
  p {{ margin: 4px 0; color: #65756d; }}
  .status {{ padding: 10px 14px; border-radius: 999px; font-size: 12px; font-weight: 700;
    color: #0d3d2c; background: #ccf06a; }}
  .summary {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 12px; margin: 22px 0; }}
  .metric {{ padding: 16px; border-radius: 12px; background: #f8f6f0; }}
  .metric span {{ display: block; color: #65756d; font-size: 11px; text-transform: uppercase;
    letter-spacing: 1px; }}
  .metric strong {{ display: block; margin-top: 8px; font-size: 20px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  th {{ padding: 11px 8px; text-align: left; color: #fff; background: #0d3d2c; }}
  td {{ padding: 11px 8px; border-bottom: 1px solid #d9ddd5; vertical-align: top; }}
  tr:nth-child(even) td {{ background: #f8f6f0; }}
  small {{ color: #65756d; }}
  .override {{ color: #815500; background: #fff0c8; border-radius: 999px; padding: 2px 6px; }}
  .footer {{ display: flex; justify-content: space-between; margin-top: 22px; color: #65756d;
    font-size: 11px; }}
  .print {{ position: fixed; top: 20px; right: 20px; border: 0; border-radius: 10px; padding: 12px 18px;
    color: white; background: #165c43; font-weight: 700; cursor: pointer; }}
  @media print {{ body {{ padding: 0; background: white; }} .report {{ border: 0; padding: 0; }}
    .print {{ display: none; }} @page {{ size: A4 landscape; margin: 10mm; }} }}
</style>
</head>
<body>
<button class="print" onclick="window.print()">Print / Save as PDF</button>
<main class="report">
  <header>
    <div>
      <div class="brand">PayYard</div>
      <h1>{escape(report.cycle.name)} payroll</h1>
      <p>{escape(report.cycle.company.legal_name)} · {report.cycle.period_start} to {report.cycle.period_end}</p>
    </div>
    <div class="status">{status_note}</div>
  </header>
  <section class="summary">
    <div class="metric"><span>Workers</span><strong>{report.totals["worker_count"]}</strong></div>
    <div class="metric"><span>Gross payroll</span><strong>{_aed(report.totals["gross_pay"])}</strong></div>
    <div class="metric"><span>Total deductions</span><strong>{_aed(report.totals["absence_deductions"] + report.totals["advance_deductions"] + report.totals["other_deductions"])}</strong></div>
    <div class="metric"><span>Final net wage</span><strong>{_aed(report.totals["final_net_pay"])}</strong></div>
    <div class="metric"><span>Manual overrides</span><strong>{report.totals["manual_override_count"]}</strong></div>
  </section>
  <table>
    <thead><tr><th>#</th><th>Code</th><th>Worker</th><th>Contract basic</th><th>Regular</th>
      <th>Overtime</th><th>Allowances</th><th>Deductions</th><th>Calculated net</th><th>Final net</th></tr></thead>
    <tbody>{''.join(rows)}</tbody>
  </table>
  <div class="footer"><span>Generated by PayYard · Cycle version {report.cycle.version}</span>
    <span>Currency: {escape(report.cycle.company.currency)}</span></div>
</main>
</body>
</html>"""


def build_pdf_report(report):
    output = io.BytesIO()
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=20,
        leading=24,
        textColor=colors.HexColor(f"#{DARK_GREEN}"),
        alignment=0,
    )
    right_style = ParagraphStyle(
        "Right", parent=styles["BodyText"], alignment=TA_RIGHT, fontSize=8
    )
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"{report.cycle.name} payroll report",
        author="PayYard",
    )
    story = [
        Table(
            [
                [
                    Paragraph("PayYard", title_style),
                    Paragraph(
                        f"<b>{escape(report.cycle.name)}</b><br/>"
                        f"{report.cycle.period_start} to {report.cycle.period_end}<br/>"
                        f"Status: {escape(report.cycle.get_status_display())}",
                        right_style,
                    ),
                ]
            ],
            colWidths=[130 * mm, 130 * mm],
        ),
        Spacer(1, 5 * mm),
    ]
    summary = [
        ["Workers", "Gross payroll", "Deductions", "Final net wage", "Overrides"],
        [
            report.totals["worker_count"],
            _aed(report.totals["gross_pay"]),
            _aed(
                report.totals["absence_deductions"]
                + report.totals["advance_deductions"]
                + report.totals["other_deductions"]
            ),
            _aed(report.totals["final_net_pay"]),
            report.totals["manual_override_count"],
        ],
    ]
    summary_table = Table(summary, colWidths=[52 * mm] * 5)
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{DARK_GREEN}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor(f"#{PAPER}")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{LINE}")),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 5 * mm)])
    rows = [
        [
            "#",
            "Code",
            "Worker",
            "Basic",
            "Regular",
            "OT",
            "Allowances",
            "Deductions",
            "Calculated",
            "Final net",
        ]
    ]
    for index, line in enumerate(report.lines, start=1):
        deductions = line.absence_deductions + line.advance_deductions + line.other_deductions
        worker_name = line.worker.full_name
        if line.manual_net_pay is not None:
            worker_name += " *"
        rows.append(
            [
                index,
                line.worker.worker_code,
                worker_name,
                f"{line.contract_basic:,.2f}",
                f"{line.regular_pay:,.2f}",
                f"{line.overtime_pay:,.2f}",
                f"{line.allowances + line.other_earnings:,.2f}",
                f"{deductions:,.2f}",
                f"{line.calculated_net_pay:,.2f}",
                f"{line.net_pay:,.2f}",
            ]
        )
    rows.append(
        [
            "",
            "",
            "TOTAL",
            f"{report.totals['contract_basic']:,.2f}",
            f"{report.totals['regular_pay']:,.2f}",
            f"{report.totals['overtime_pay']:,.2f}",
            f"{report.totals['allowances'] + report.totals['other_earnings']:,.2f}",
            f"{report.totals['absence_deductions'] + report.totals['advance_deductions'] + report.totals['other_deductions']:,.2f}",
            f"{report.totals['calculated_net_pay']:,.2f}",
            f"{report.totals['final_net_pay']:,.2f}",
        ]
    )
    payroll_table = Table(
        rows,
        repeatRows=1,
        colWidths=[8 * mm, 18 * mm, 41 * mm, 25 * mm, 25 * mm, 20 * mm, 25 * mm, 25 * mm, 27 * mm, 27 * mm],
    )
    payroll_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{DARK_GREEN}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(f"#{LIGHT_GREEN}")),
                ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{LINE}")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor(f"#{PAPER}")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(payroll_table)
    if report.totals["manual_override_count"]:
        story.extend(
            [
                Spacer(1, 3 * mm),
                Paragraph(
                    "* Final wage includes an authorized manual override. "
                    "See the PayYard audit log for actor, timestamp, and reason.",
                    styles["BodyText"],
                ),
            ]
        )
    document.build(story)
    return output.getvalue(), _filename(report, "pdf")


def build_excel_report(report):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:F2")
    title = summary["A1"]
    title.value = f"PayYard | {report.cycle.name} Payroll Report"
    title.font = Font(size=20, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=DARK_GREEN)
    title.alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 26
    summary["A4"] = "Company"
    summary["B4"] = report.cycle.company.legal_name
    summary["A5"] = "Period"
    summary["B5"] = f"{report.cycle.period_start} to {report.cycle.period_end}"
    summary["A6"] = "Cycle status"
    summary["B6"] = report.cycle.get_status_display()
    summary["A7"] = "Cycle version"
    summary["B7"] = report.cycle.version
    metrics = [
        ("Workers", report.totals["worker_count"]),
        ("Gross payroll", report.totals["gross_pay"]),
        (
            "Total deductions",
            report.totals["absence_deductions"]
            + report.totals["advance_deductions"]
            + report.totals["other_deductions"],
        ),
        ("Calculated net", report.totals["calculated_net_pay"]),
        ("Final net wage", report.totals["final_net_pay"]),
        ("Manual overrides", report.totals["manual_override_count"]),
    ]
    for column, (label, value) in enumerate(metrics, start=1):
        label_cell = summary.cell(row=10, column=column, value=label)
        value_cell = summary.cell(row=11, column=column, value=value)
        label_cell.fill = PatternFill("solid", fgColor=DARK_GREEN)
        label_cell.font = Font(color="FFFFFF", bold=True)
        value_cell.fill = PatternFill("solid", fgColor=PAPER)
        value_cell.font = Font(size=14, bold=True, color=DARK_GREEN)
        for cell in (label_cell, value_cell):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(value, Decimal):
            value_cell.number_format = '"AED" #,##0.00'
    for column in range(1, 7):
        summary.column_dimensions[get_column_letter(column)].width = 22

    sheet = workbook.create_sheet("Worker wages")
    sheet.sheet_view.showGridLines = False
    headers = [
        "#",
        "Worker code",
        "Worker name",
        "Job title",
        "Wage type",
        "Contract basic",
        "Regular pay",
        "Overtime",
        "Allowances",
        "Other earnings",
        "Absence deductions",
        "Advance deductions",
        "Other deductions",
        "Gross pay",
        "Calculated net",
        "Manual override",
        "Final net wage",
        "Override reason",
    ]
    sheet.append(headers)
    for index, line in enumerate(report.lines, start=1):
        sheet.append(
            [
                index,
                line.worker.worker_code,
                line.worker.full_name,
                line.worker.job_title,
                line.worker.get_wage_type_display(),
                line.contract_basic,
                line.regular_pay,
                line.overtime_pay,
                line.allowances,
                line.other_earnings,
                line.absence_deductions,
                line.advance_deductions,
                line.other_deductions,
                line.gross_pay,
                line.calculated_net_pay,
                line.manual_net_pay,
                line.net_pay,
                line.manual_override_reason,
            ]
        )
    total_row = sheet.max_row + 1
    sheet.cell(total_row, 3, "TOTAL")
    for column in range(6, 18):
        letter = get_column_letter(column)
        sheet.cell(total_row, column, f"=SUM({letter}2:{letter}{total_row - 1})")
    header_fill = PatternFill("solid", fgColor=DARK_GREEN)
    total_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    thin = Side(style="thin", color=LINE)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    for cell in sheet[total_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True, color=DARK_GREEN)
        cell.border = Border(top=thin)
    for row in sheet.iter_rows(min_row=2, max_row=total_row, min_col=6, max_col=17):
        for cell in row:
            cell.number_format = '"AED" #,##0.00'
    for row_number in range(2, total_row):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = PatternFill("solid", fgColor=PAPER)
    widths = [7, 16, 26, 22, 15, 18, 18, 15, 17, 17, 20, 20, 18, 18, 19, 18, 18, 38]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:R{total_row - 1}"
    sheet.row_dimensions[1].height = 28
    sheet.print_title_rows = "1:1"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue(), _filename(report, "xlsx")
